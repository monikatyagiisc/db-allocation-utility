from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.logging_config import get_logger
from app.models import DatabaseRecord

logger = get_logger("app.excel")

MAX_HEADER_COLS = 16

EXCEL_HEADERS = [
    "S.N.",
    "Type",
    "Database",
    "# of CICS Trns",
    "Prod Mirror",
    "Release",
    "Lifecycle",
    "Status",
    "Assignee",
    "Team",
    "Project",
    "Start Date",
    "End Date",
    "Can be released -Y/N",
    "Comments",
]

# Normalized header label -> model field
HEADER_ALIASES: dict[str, str] = {
    "s.n.": "serial_number",
    "s.no.": "serial_number",
    "sn": "serial_number",
    "type": "database_type",
    "database": "database_name",
    "of cics trns": "cics_transactions",
    "cics trns": "cics_transactions",
    "trans": "cics_transactions",
    "prod mirror": "prod_mirror",
    "release": "release",
    "lifecycle": "lifecycle",
    "status": "status",
    "assignee": "assignee",
    "team": "team",
    "project": "project",
    "start date": "start_date",
    "end date": "end_date",
    "can be released -y/n": "can_be_released",
    "can be released": "can_be_released",
    "comments": "comments",
}


def _normalize_header(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip().lower()
    if not s or s.startswith("db with"):
        return None
    if s.startswith("#"):
        s = s.lstrip("#").strip()
    return s


def _parse_header_map(header_row: tuple) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for idx in range(min(len(header_row), MAX_HEADER_COLS)):
        norm = _normalize_header(header_row[idx])
        if not norm:
            continue
        field = HEADER_ALIASES.get(norm)
        if field and field not in col_map:
            col_map[field] = idx
    return col_map


def _get_cell(row: tuple, col_map: dict[str, int], field: str):
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _cell_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _cell_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _cell_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def row_to_record(row: tuple, col_map: dict[str, int], database_type: str) -> DatabaseRecord | None:
    database_name = _cell_str(_get_cell(row, col_map, "database_name"))
    if not database_name:
        return None
    return DatabaseRecord(
        serial_number=_cell_int(_get_cell(row, col_map, "serial_number")),
        database_type=database_type,
        database_name=database_name,
        cics_transactions=_cell_int(_get_cell(row, col_map, "cics_transactions")),
        prod_mirror=_cell_str(_get_cell(row, col_map, "prod_mirror")),
        release=_cell_str(_get_cell(row, col_map, "release")),
        lifecycle=_cell_str(_get_cell(row, col_map, "lifecycle")),
        status=_cell_str(_get_cell(row, col_map, "status")),
        assignee=_cell_str(_get_cell(row, col_map, "assignee")),
        team=_cell_str(_get_cell(row, col_map, "team")),
        project=_cell_str(_get_cell(row, col_map, "project")),
        start_date=_cell_date(_get_cell(row, col_map, "start_date")),
        end_date=_cell_date(_get_cell(row, col_map, "end_date")),
        can_be_released=_cell_str(_get_cell(row, col_map, "can_be_released")),
        comments=_cell_str(_get_cell(row, col_map, "comments")),
    )


def _import_sheet(ws, database_type: str, db: Session) -> tuple[int, int]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return 0, 0

    col_map = _parse_header_map(header_row)
    if "database_name" not in col_map:
        logger.warning("Sheet '%s' skipped: no Database column in header", database_type)
        return 0, 0

    imported = 0
    skipped = 0
    for row in rows_iter:
        if not row or all(c is None or str(c).strip() == "" for c in row[:MAX_HEADER_COLS]):
            skipped += 1
            continue
        record = row_to_record(row, col_map, database_type)
        if record is None:
            skipped += 1
            continue
        db.add(record)
        imported += 1
    return imported, skipped


def import_excel(db: Session, file_bytes: bytes, replace: bool = False) -> tuple[int, int, list[str]]:
    logger.info("Parsing Excel file size=%s bytes replace=%s", len(file_bytes), replace)
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        logger.exception("Failed to open Excel workbook")
        raise

    sheet_names = list(wb.sheetnames)
    logger.info("Workbook sheets: %s", sheet_names)

    if replace:
        deleted = db.query(DatabaseRecord).delete()
        db.commit()
        logger.info("Replaced existing records deleted=%s", deleted)

    total_imported = 0
    total_skipped = 0
    sheets_processed: list[str] = []

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        imported, skipped = _import_sheet(ws, sheet_name.strip(), db)
        if imported > 0 or skipped > 0:
            sheets_processed.append(sheet_name)
            logger.info("Sheet '%s': imported=%s skipped=%s", sheet_name, imported, skipped)
        total_imported += imported
        total_skipped += skipped

    wb.close()
    db.commit()
    logger.info(
        "Excel import finished imported=%s skipped=%s sheets=%s",
        total_imported,
        total_skipped,
        sheets_processed,
    )
    return total_imported, total_skipped, sheets_processed


def record_to_row(record: DatabaseRecord) -> list:
    return [
        record.serial_number,
        record.database_type,
        record.database_name,
        record.cics_transactions,
        record.prod_mirror,
        record.release,
        record.lifecycle,
        record.status,
        record.assignee,
        record.team,
        record.project,
        record.start_date,
        record.end_date,
        record.can_be_released,
        record.comments,
    ]


def export_excel(db: Session) -> bytes:
    records = (
        db.query(DatabaseRecord)
        .order_by(DatabaseRecord.database_type, DatabaseRecord.serial_number, DatabaseRecord.id)
        .all()
    )
    logger.info("Building Excel export record_count=%s", len(records))

    by_type: dict[str, list[DatabaseRecord]] = {}
    for record in records:
        sheet = record.database_type or "Unassigned"
        by_type.setdefault(sheet, []).append(record)

    wb = Workbook()
    wb.remove(wb.active)

    if not by_type:
        ws = wb.create_sheet("Sheet1")
        ws.append(EXCEL_HEADERS)
    else:
        for sheet_name, sheet_records in sorted(by_type.items()):
            safe_name = sheet_name[:31] if sheet_name else "Unassigned"
            ws = wb.create_sheet(safe_name)
            ws.append(EXCEL_HEADERS)
            for record in sheet_records:
                ws.append(record_to_row(record))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
